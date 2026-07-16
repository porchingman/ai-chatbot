import io
import httpx
import os
from pypdf import PdfReader
from docx import Document
import openpyxl
from fastapi import HTTPException
from google import genai
from google.genai import types

from app.config import settings
from app.database import get_supabase
from app.ai import get_ai_client, EMBEDDING_MODEL

ai_client = get_ai_client()

class RAGService:
    @staticmethod
    def extract_text_from_docx(file_stream) -> list[dict]:
        """Word(.docx) 파일에서 텍스트를 페이지 개념 대신 일정 문단 단위로 추출"""
        doc = Document(file_stream)
        text_list = []
        current_text = []
        
        for para in doc.paragraphs:
            if para.text.strip():
                current_text.append(para.text.strip())
            # 500글자 정도 모이면 가상의 페이지(단락 블록)로 분할
            if len("\n".join(current_text)) > 500:
                text_list.append("\n".join(current_text))
                current_text = []
                
        if current_text:
            text_list.append("\n".join(current_text))
            
        return [{"page_no": i, "text": text} for i, text in enumerate(text_list, start=1)]

    @staticmethod
    def extract_text_from_xlsx(file_stream) -> list[dict]:
        """Excel(.xlsx) 파일에서 행 데이터를 텍스트로 보존하며 시트 단위 추출"""
        wb = openpyxl.load_workbook(file_stream, data_only=True)
        pages_content = []
        
        for sheet_idx, sheet_name in enumerate(wb.sheetnames, start=1):
            sheet = wb[sheet_name]
            sheet_text = [f"--- 시트명: {sheet_name} ---"]
            
            for row in sheet.iter_rows(values_only=True):
                # 공백 셀 제외하고 한 줄의 텍스트 라인 조립
                row_text = ", ".join([str(cell).strip() for cell in row if cell is not None])
                if row_text.strip():
                    sheet_text.append(row_text)
                    
            if len(sheet_text) > 1:
                pages_content.append({
                    "page_no": sheet_idx,
                    "text": "\n".join(sheet_text)
                })
        return pages_content

    @classmethod
    async def download_and_extract_text(cls, url: str) -> list[dict]:
        """
        [최적화] URL 주소의 확장자를 판별하여 동적으로 적절한 파서 엔진을 가동합니다.
        (PDF, DOCX, XLSX 표준 3종 포맷만 집중 대응)
        """
        # URL에서 순수 파일 확장자 추출 및 소문자 변환
        pure_path = url.split("?")[0].lower()
        _, ext = os.path.splitext(pure_path)

        async with httpx.AsyncClient(timeout=45.0) as client:
            try:
                response = await client.get(url)
                response.raise_for_status()
            except Exception as e:
                raise HTTPException(status_code=400, detail=f"파일 다운로드 실패: {str(e)}")

        file_stream = io.BytesIO(response.content)

        try:
            if ext == ".pdf":
                reader = PdfReader(file_stream)
                pages_content = []
                for page_num, page in enumerate(reader.pages, start=1):
                    text = page.extract_text()
                    if text and text.strip():
                        pages_content.append({"page_no": page_num, "text": text.strip()})
                return pages_content
                
            elif ext == ".docx":
                return cls.extract_text_from_docx(file_stream)
                
            elif ext == ".xlsx":
                return cls.extract_text_from_xlsx(file_stream)
                
            else:
                # 한글 파일 차단 및 상용 확장자 제한 메시지 출력
                raise HTTPException(status_code=415, detail=f"지원하지 않는 확장자({ext})입니다. (PDF, DOCX, XLSX 파일만 등록 가능합니다)")
                
        except HTTPException:
            raise
        except Exception as e:
            raise HTTPException(status_code=422, detail=f"문서 구조 파싱 실패: {str(e)}")

    @classmethod
    async def process_and_save_document(cls, company_code: int, req) -> dict:
        """문서를 마스터(knowledge)와 청크(knowledge_data)로 분리하여 저장하는 파이프라인"""
        supabase = get_supabase()

        # 설계 원칙: 기존 동일 식별자 문서 선행 일괄 제거 (CASCADE 연동 소멸)
        supabase.table("knowledge").delete()\
            .eq("company_code", company_code)\
            .eq("source_type", req.source_type)\
            .eq("source_code", req.source_code).execute()

        # 파일 형식별 다운로드 및 자동 텍스트 추출 호출
        pages = await cls.download_and_extract_text(str(req.source_url))

        # 마스터 테이블 저장 (content 컬럼 비우기 완전 적용)
        master_data = {
            "company_code": company_code,
            "source_type": req.source_type,
            "source_code": req.source_code,
            "source_url": str(req.source_url),
            "title": req.title,
            "total_token": 0 
        }
        master_insert = supabase.table("knowledge").insert(master_data).execute()
        if not master_insert.data:
            raise HTTPException(status_code=500, detail="마스터 문서 정보 생성 실패")
        
        # Supabase SDK 반환 리스트에서 첫 번째 객체의 code 추출
        knowledge_code = master_insert.data[0]["code"]

        # 청킹 스플리터 가동
        from langchain_text_splitters import RecursiveCharacterTextSplitter
        text_splitter = RecursiveCharacterTextSplitter(chunk_size=1000, chunk_overlap=200)

        chunks_payload = []
        chunk_no = 1

        for page_data in pages:
            split_texts = text_splitter.split_text(page_data["text"])
            for text_chunk in split_texts:
                chunks_payload.append({
                    "knowledge_code": knowledge_code,
                    "company_code": company_code,
                    "content": text_chunk, 
                    "chunk_no": chunk_no,
                    "page_no": page_data["page_no"],
                    "token": len(text_chunk)
                })
                chunk_no += 1

        # Gemini 임베딩 생성 (768차원 최신 제한 옵션 일치)
        try:
            texts_to_embed = [c["content"] for c in chunks_payload]
            embed_response = ai_client.models.embed_content(
                model=EMBEDDING_MODEL,
                contents=texts_to_embed,
                config=types.EmbedContentConfig(output_dimensionality=768)
            )
            for idx, embedding_data in enumerate(embed_response.embeddings):
                chunks_payload[idx]["embedding"] = embedding_data.values
                chunks_payload[idx]["embedding_model"] = EMBEDDING_MODEL
        except Exception as e:
            # 실패 시 트랜잭션 복구 안전장치
            supabase.table("knowledge").delete().eq("code", knowledge_code).execute()
            raise HTTPException(status_code=500, detail=f"Gemini 임베딩 생성 오류: {str(e)}")

        if chunks_payload:
            supabase.table("knowledge_data").insert(chunks_payload).execute()

        total_calculated_tokens = sum([c["token"] for c in chunks_payload])
        supabase.table("knowledge").update({"total_token": total_calculated_tokens}).eq("code", knowledge_code).execute()

        return {
            "knowledge_code": knowledge_code,
            "total_chunks": len(chunks_payload)
        }
